"""
Обработчики разделов категории: медиа, описание, цены, отзывы
"""
from telebot import types
from loader import bot
from database.database import db
from config import ADMIN_ID
from utils import escape_html, safe_answer_callback


# ═══════════════════════════════════════════════════════════════
# МЕДИА
# ═══════════════════════════════════════════════════════════════

@bot.callback_query_handler(func=lambda call: call.data.startswith("category_media_"))
def handle_category_media(call):
    """Управление медиа категории"""
    category_id = int(call.data.split("_")[-1])
    
    category = db.get_category(category_id)
    if not category:
        safe_answer_callback(bot, call.id, "❌ Категория не найдена")
        return
    
    category_name = category['name']
    media = category.get('media', [])
    media_count = len(media) if isinstance(media, list) else 0
    
    text = (
        f"📷 <b>МОИ МЕДИА</b>\n"
        f"📂 Категория: {escape_html(category_name)}\n"
        "━━━━━━━━━━━━━━\n\n"
    )
    
    if media_count > 0:
        text += (
            f"📊 Загружено файлов: <b>{media_count}</b>\n\n"
            "Вы можете:\n"
            "• Просмотреть галерею (скоро)\n"
            "• Загрузить новые файлы (скоро)\n"
            "• Сгенерировать AI-изображение\n"
        )
    else:
        text += (
            "У вас пока нет медиа-файлов для этой категории.\n\n"
            "🎨 <b>AI-генерация изображений:</b>\n"
            "Создайте уникальное изображение для вашего товара/услуги с помощью Nano Banana Pro!\n\n"
            "💰 <b>Стоимость:</b> 30 токенов\n"
        )
    
    markup = types.InlineKeyboardMarkup(row_width=1)
    
    if media_count > 0:
        markup.add(
            types.InlineKeyboardButton("📷 Просмотреть галерею", callback_data=f"view_gallery_{category_id}")
        )
    
    markup.add(
        types.InlineKeyboardButton("🎨 Сгенерировать изображение", callback_data=f"gen_image_{category_id}"),
        types.InlineKeyboardButton("📤 Загрузить файлы", callback_data=f"upload_media_{category_id}"),
        types.InlineKeyboardButton("🔙 К категории", callback_data=f"open_category_{category_id}")
    )
    
    try:
        bot.edit_message_text(
            text,
            call.message.chat.id,
            call.message.message_id,
            reply_markup=markup,
            parse_mode='HTML'
        )
    except:
        bot.send_message(call.message.chat.id, text, reply_markup=markup, parse_mode='HTML')
    
    safe_answer_callback(bot, call.id)


# ═══════════════════════════════════════════════════════════════
# ОПИСАНИЕ
# ═══════════════════════════════════════════════════════════════

@bot.callback_query_handler(func=lambda call: call.data.startswith("category_description_"))
def handle_category_description(call):
    """Управление описанием категории"""
    category_id = int(call.data.split("_")[-1])
    
    category = db.get_category(category_id)
    if not category:
        safe_answer_callback(bot, call.id, "❌ Категория не найдена")
        return
    
    category_name = category['name']
    description = category.get('description', '')
    
    text = (
        f"📝 <b>ОПИСАНИЕ ПРОДУКТА/УСЛУГИ</b>\n"
        f"📂 Категория: {escape_html(category_name)}\n"
        "━━━━━━━━━━━━━━\n\n"
    )
    
    if description:
        # Показываем существующее описание
        desc_preview = description[:300]
        if len(description) > 300:
            desc_preview += "..."
        
        text += (
            f"<b>Текущее описание:</b>\n\n"
            f"<i>{escape_html(desc_preview)}</i>\n\n"
            "━━━━━━━━━━━━━━\n\n"
            "💡 <i>Это описание используется для генерации изображений</i>\n\n"
            "Вы можете:\n"
            "• Редактировать существующее описание\n"
            "• Удалить описание"
        )
        
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            types.InlineKeyboardButton("✏️ Редактировать", callback_data=f"edit_description_{category_id}"),
            types.InlineKeyboardButton("🗑 Удалить описание", callback_data=f"delete_description_{category_id}"),
            types.InlineKeyboardButton("🔙 К категории", callback_data=f"open_category_{category_id}")
        )
    else:
        # Предлагаем добавить описание
        text += (
            "У вас пока нет описания для этой категории.\n\n"
            "📝 <b>Добавьте описание</b>\n\n"
            "Описание используется для:\n"
            "• 🎨 Генерации изображений с помощью AI\n"
            "• 📱 Постов в социальных сетях\n"
            "• 🌐 Публикаций на сайте\n\n"
            "💡 <i>Опишите продукт так, как вы хотите видеть его на изображениях</i>\n\n"
            "<b>Например:</b>\n"
            "<i>\"WPC панели от компании ООО 'Дизайн-Сервис' — современное решение "
            "для отделки стен, сочетающее в себе эстетику натурального дерева и практичность "
            "композитных материалов. Наши стеновые панели WPC предст...\"</i>"
        )
        
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            types.InlineKeyboardButton("➕ Добавить описание", callback_data=f"add_description_{category_id}"),
            types.InlineKeyboardButton("🔙 К категории", callback_data=f"open_category_{category_id}")
        )
    
    try:
        bot.edit_message_text(
            text,
            call.message.chat.id,
            call.message.message_id,
            reply_markup=markup,
            parse_mode='HTML'
        )
    except:
        bot.send_message(call.message.chat.id, text, reply_markup=markup, parse_mode='HTML')
    
    safe_answer_callback(bot, call.id)


@bot.callback_query_handler(func=lambda call: call.data.startswith("add_description_"))
def handle_add_description(call):
    """Добавление описания категории"""
    category_id = int(call.data.split("_")[-1])
    user_id = call.from_user.id
    
    # Сохраняем состояние ожидания описания
    from handlers.categories import category_creation_state
    category_creation_state[user_id] = {
        'action': 'add_description',
        'category_id': category_id
    }
    
    text = (
        "📝 <b>ДОБАВЛЕНИЕ ОПИСАНИЯ</b>\n"
        "━━━━━━━━━━━━━━\n\n"
        "✍️ Напишите описание продукта/услуги.\n\n"
        "💡 <b>Советы:</b>\n"
        "• Опишите товар так, как хотите видеть на изображениях\n"
        "• Укажите ключевые особенности и преимущества\n"
        "• Используйте понятный язык\n\n"
        "📤 Отправьте текст описания:"
    )
    
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("❌ Отмена", callback_data=f"category_description_{category_id}")
    )
    
    try:
        bot.edit_message_text(
            text,
            call.message.chat.id,
            call.message.message_id,
            reply_markup=markup,
            parse_mode='HTML'
        )
    except:
        bot.send_message(call.message.chat.id, text, reply_markup=markup, parse_mode='HTML')
    
    safe_answer_callback(bot, call.id)


@bot.callback_query_handler(func=lambda call: call.data.startswith("edit_description_"))
def handle_edit_description_start(call):
    """Начало редактирования описания"""
    category_id = int(call.data.split("_")[-1])
    user_id = call.from_user.id
    
    category = db.get_category(category_id)
    if not category:
        safe_answer_callback(bot, call.id, "❌ Категория не найдена")
        return
    
    current_description = category.get('description', '')
    
    # Сохраняем состояние
    from handlers.categories import category_creation_state
    category_creation_state[user_id] = {
        'action': 'edit_description',
        'category_id': category_id
    }
    
    text = (
        "✏️ <b>РЕДАКТИРОВАНИЕ ОПИСАНИЯ</b>\n"
        "━━━━━━━━━━━━━━\n\n"
        f"<b>Текущее описание:</b>\n\n"
        f"<i>{escape_html(current_description)}</i>\n\n"
        "━━━━━━━━━━━━━━\n\n"
        "💡 <b>Как писать описание для AI-генерации:</b>\n\n"
        "Напишите <b>3-5 разных вариантов</b> описаний через <b>запятую</b>.\n"
        "При генерации изображений AI будет случайно выбирать <b>1-2 фразы</b>.\n\n"
        "⚠️ <b>ВАЖНО:</b> Каждая фраза должна быть <b>РАЗНОЙ</b>!\n"
        "Не копируйте одно и то же много раз.\n\n"
        "<b>Пример:</b>\n"
        "<code>luxury интерьеры со стеновыми панелями большого формата, "
        "классический интерьер с деревянными панелями, "
        "современный минимализм с WPC панелями, "
        "премиальный дизайн с текстурными панелями</code>\n\n"
        "━━━━━━━━━━━━━━\n\n"
        "📤 Отправьте новый текст описания:"
    )
    
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("❌ Отмена", callback_data=f"category_description_{category_id}")
    )
    
    try:
        bot.edit_message_text(
            text,
            call.message.chat.id,
            call.message.message_id,
            reply_markup=markup,
            parse_mode='HTML'
        )
    except:
        bot.send_message(call.message.chat.id, text, reply_markup=markup, parse_mode='HTML')
    
    safe_answer_callback(bot, call.id)


@bot.callback_query_handler(func=lambda call: call.data.startswith("delete_description_"))
def handle_delete_description(call):
    """Удаление описания"""
    category_id = int(call.data.split("_")[-1])
    
    # Удаляем описание
    try:
        db.cursor.execute("""
            UPDATE categories 
            SET description = NULL
            WHERE id = %s
        """, (category_id,))
        db.conn.commit()
        
        text = (
            "✅ <b>ОПИСАНИЕ УДАЛЕНО</b>\n"
            "━━━━━━━━━━━━━━\n\n"
            "Описание категории успешно удалено."
        )
        
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("🔙 К категории", callback_data=f"open_category_{category_id}")
        )
        
        bot.edit_message_text(
            text,
            call.message.chat.id,
            call.message.message_id,
            reply_markup=markup,
            parse_mode='HTML'
        )
        
        safe_answer_callback(bot, call.id, "✅ Удалено")
        
    except Exception as e:
        print(f"❌ Ошибка удаления описания: {e}")
        safe_answer_callback(bot, call.id, "❌ Ошибка", show_alert=True)


# Обработчик текстового сообщения для добавления/редактирования описания
@bot.message_handler(func=lambda message: message.from_user.id in __import__('handlers.categories', fromlist=['category_creation_state']).category_creation_state 
                     and __import__('handlers.categories', fromlist=['category_creation_state']).category_creation_state[message.from_user.id].get('action') in ['add_description', 'edit_description'])
def handle_description_text(message):
    """Обработка текста описания"""
    user_id = message.from_user.id
    from handlers.categories import category_creation_state
    
    state = category_creation_state.get(user_id)
    if not state:
        return
    
    category_id = state['category_id']
    action = state['action']
    description_text = message.text.strip()
    
    if len(description_text) < 10:
        bot.send_message(
            message.chat.id,
            "❌ Описание слишком короткое. Минимум 10 символов."
        )
        return
    
    try:
        # Сохраняем описание
        db.cursor.execute("""
            UPDATE categories 
            SET description = %s
            WHERE id = %s
        """, (description_text, category_id))
        db.conn.commit()
        
        # Очищаем состояние
        del category_creation_state[user_id]
        
        action_text = "добавлено" if action == 'add_description' else "обновлено"
        
        text = (
            f"✅ <b>ОПИСАНИЕ {action_text.upper()}</b>\n"
            "━━━━━━━━━━━━━━\n\n"
            f"<b>Ваше описание:</b>\n\n"
            f"<i>{escape_html(description_text[:300])}</i>\n"
        )
        
        if len(description_text) > 300:
            text += "\n<i>... (текст сокращен)</i>"
        
        text += "\n\n💡 <i>Это описание будет использоваться для генерации изображений</i>"
        
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("🔙 К категории", callback_data=f"open_category_{category_id}")
        )
        
        bot.send_message(message.chat.id, text, reply_markup=markup, parse_mode='HTML')
        
    except Exception as e:
        print(f"❌ Ошибка сохранения описания: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка сохранения: {e}")


# ═══════════════════════════════════════════════════════════════
# ЦЕНЫ
# ═══════════════════════════════════════════════════════════════

@bot.callback_query_handler(func=lambda call: call.data.startswith("category_prices_"))
def handle_category_prices(call):
    """Управление ценами категории"""
    category_id = int(call.data.split("_")[-1])
    
    category = db.get_category(category_id)
    if not category:
        safe_answer_callback(bot, call.id, "❌ Категория не найдена")
        return
    
    category_name = category['name']
    prices = category.get('prices', {})
    has_prices = bool(prices)
    
    text = (
        f"💰 <b>ЦЕНЫ</b>\n"
        f"📂 Категория: {escape_html(category_name)}\n"
        "━━━━━━━━━━━━━━\n\n"
    )
    
    if has_prices:
        text += (
            "✅ Прайс-лист загружен!\n\n"
            "Вы можете:\n"
            "• Скачать текущий прайс\n"
            "• Загрузить обновленный Excel файл\n"
            "• Удалить прайс-лист\n\n"
            "<i>Функция управления ценами в разработке</i>"
        )
        
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            types.InlineKeyboardButton("📥 Скачать текущий прайс", callback_data=f"download_current_price_{category_id}"),
            types.InlineKeyboardButton("📤 Загрузить обновленный", callback_data=f"upload_price_file_{category_id}"),
            types.InlineKeyboardButton("🗑 Удалить прайс-лист", callback_data=f"delete_price_{category_id}"),
            types.InlineKeyboardButton("🔙 К категории", callback_data=f"open_category_{category_id}")
        )
    else:
        text += (
            "У вас пока нет прайс-листа.\n\n"
            "📊 <b>Загрузка прайса из Excel:</b>\n\n"
            "1️⃣ Скачайте шаблон Excel\n"
            "2️⃣ Заполните товары и цены\n"
            "3️⃣ Загрузите обратно в бот\n\n"
            "Прайс-лист будет использоваться при генерации описаний и контента!"
        )
        
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            types.InlineKeyboardButton("📥 Скачать шаблон Excel", callback_data=f"download_price_template_{category_id}"),
            types.InlineKeyboardButton("📤 Загрузить прайс", callback_data=f"upload_price_file_{category_id}"),
            types.InlineKeyboardButton("🔙 К категории", callback_data=f"open_category_{category_id}")
        )
    
    try:
        bot.edit_message_text(
            text,
            call.message.chat.id,
            call.message.message_id,
            reply_markup=markup,
            parse_mode='HTML'
        )
    except:
        bot.send_message(call.message.chat.id, text, reply_markup=markup, parse_mode='HTML')
    
    safe_answer_callback(bot, call.id)


# ═══════════════════════════════════════════════════════════════
# ОТЗЫВЫ
# ═══════════════════════════════════════════════════════════════

@bot.callback_query_handler(func=lambda call: call.data.startswith("category_reviews_"))
def handle_category_reviews(call):
    """Управление отзывами категории"""
    category_id = int(call.data.split("_")[-1])
    
    category = db.get_category(category_id)
    if not category:
        safe_answer_callback(bot, call.id, "❌ Категория не найдена")
        return
    
    category_name = category['name']
    reviews = category.get('reviews', [])
    reviews_count = len(reviews) if isinstance(reviews, list) else 0
    
    text = (
        f"⭐️ <b>ОТЗЫВЫ</b>\n"
        f"📂 Категория: {escape_html(category_name)}\n"
        "━━━━━━━━━━━━━━\n\n"
    )
    
    if reviews_count > 0:
        text += (
            f"📊 Всего отзывов: <b>{reviews_count}</b>\n\n"
            "Отзывы сохранены и готовы к использованию!\n\n"
            "Вы можете:\n"
            "• Просмотреть все отзывы\n"
            "• Добавить ещё отзывов\n"
            "• Использовать их в контенте"
        )
        
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            types.InlineKeyboardButton("📋 Посмотреть все отзывы", callback_data=f"view_all_reviews_{category_id}"),
            types.InlineKeyboardButton("➕ Добавить ещё отзывов", callback_data=f"gen_reviews_{category_id}"),
            types.InlineKeyboardButton("🔙 К категории", callback_data=f"open_category_{category_id}")
        )
    else:
        text += (
            "У вас пока нет отзывов.\n\n"
            "⭐️ <b>Отзывы опциональны</b>\n\n"
            "Отзывы нужны только если вы планируете:\n"
            "• Создавать сайт с отзывами\n"
            "• Использовать их в контенте\n"
            "• Показывать социальное доказательство\n\n"
            "💡 <b>AI-генерация отзывов:</b>\n"
            "Создайте реалистичные отзывы клиентов с помощью Claude AI!\n\n"
            "💰 <b>Стоимость:</b>\n"
            "• 3 отзыва — 30 токенов\n"
            "• 5 отзывов — 50 токенов\n"
            "• 10 отзывов — 100 токенов"
        )
        
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            types.InlineKeyboardButton("🤖 Сгенерировать отзывы", callback_data=f"gen_reviews_{category_id}"),
            types.InlineKeyboardButton("🔙 К категории", callback_data=f"open_category_{category_id}")
        )
    
    try:
        bot.edit_message_text(
            text,
            call.message.chat.id,
            call.message.message_id,
            reply_markup=markup,
            parse_mode='HTML'
        )
    except:
        bot.send_message(call.message.chat.id, text, reply_markup=markup, parse_mode='HTML')
    
    safe_answer_callback(bot, call.id)


# ═══════════════════════════════════════════════════════════════
# ОБЩАЯ ЗАГЛУШКА "СКОРО"
# ═══════════════════════════════════════════════════════════════

@bot.callback_query_handler(func=lambda call: call.data == "coming_soon")
def handle_coming_soon(call):
    """Заглушка для функций в разработке"""
    safe_answer_callback(bot, call.id, "🚧 Функция в разработке", show_alert=True)


# ═══════════════════════════════════════════════════════════════
# ГЕНЕРАЦИЯ ОПИСАНИЯ
# ═══════════════════════════════════════════════════════════════

@bot.callback_query_handler(func=lambda call: call.data.startswith("gen_desc_"))
def handle_generate_description(call):
    """Генерация описания категории с помощью AI"""
    from ai.text_generator import generate_product_description
    
    category_id = int(call.data.split("_")[-1])
    user_id = call.from_user.id
    
    # Получаем категорию
    category = db.get_category(category_id)
    if not category:
        safe_answer_callback(bot, call.id, "❌ Категория не найдена")
        return
    
    # Получаем бота
    bot_data = db.get_bot(category['bot_id'])
    if not bot_data:
        safe_answer_callback(bot, call.id, "❌ Бот не найден")
        return
    
    # Проверяем баланс (20 токенов за описание)
    user = db.get_user(user_id)
    tokens = user.get('tokens', 0)
    cost = 20
    
    if str(user_id) != str(ADMIN_ID) and tokens < cost:
        bot.answer_callback_query(
            call.id,
            f"❌ Недостаточно токенов. Нужно: {cost} 💎",
            show_alert=True
        )
        return
    
    # Показываем прогресс
    try:
        bot.edit_message_text(
            "🤖 <b>AI генерирует описание...</b>\n\n"
            "⏳ Это может занять 10-20 секунд...",
            call.message.chat.id,
            call.message.message_id,
            parse_mode='HTML'
        )
    except:
        pass
    
    safe_answer_callback(bot, call.id, "🚀 Начинаем генерацию...")
    
    # Собираем данные для генерации
    category_name = category['name']
    company_data = bot_data.get('company_data', {})
    keywords = category.get('keywords', [])
    
    # Формируем характеристики и преимущества
    features = f"Категория: {category_name}"
    if keywords:
        features += f"\nКлючевые запросы: {', '.join(keywords[:5])}"
    
    benefits = "Высокое качество, профессиональный подход"
    if company_data.get('company_name'):
        benefits = f"От компании {company_data['company_name']} - {benefits}"
    
    target_audience = company_data.get('target_audience', 'Широкая аудитория')
    
    # Генерируем описание
    result = generate_product_description(
        product_name=category_name,
        category=category_name,
        features=features,
        benefits=benefits,
        target_audience=target_audience,
        tone='professional',
        length='medium'
    )
    
    if not result['success']:
        bot.send_message(
            call.message.chat.id,
            f"❌ <b>Ошибка генерации</b>\n\n{result.get('error', 'Неизвестная ошибка')}",
            parse_mode='HTML'
        )
        return
    
    description = result['text']
    word_count = result.get('word_count', 0)
    
    # Списываем токены
    if str(user_id) != str(ADMIN_ID):
        db.update_tokens(user_id, -cost)
        
        # Логируем расход
        db.cursor.execute("""
            INSERT INTO token_expenses (user_id, amount, action, category_id, created_at)
            VALUES (%s, %s, %s, %s, NOW())
        """, (user_id, cost, 'text_generation', category_id))
        db.conn.commit()
    
    # Сохраняем описание
    db.cursor.execute("""
        UPDATE categories 
        SET description = %s
        WHERE id = %s
    """, (description, category_id))
    db.conn.commit()
    
    # Показываем результат
    preview = description[:300] + "..." if len(description) > 300 else description
    
    text = (
        "✅ <b>ОПИСАНИЕ СГЕНЕРИРОВАНО!</b>\n"
        "━━━━━━━━━━━━━━\n\n"
        f"📂 Категория: <b>{escape_html(category_name)}</b>\n"
        f"📝 Слов: <b>{word_count}</b>\n"
        f"💎 Списано токенов: <b>{cost}</b>\n\n"
        "━━━━━━━━━━━━━━\n\n"
        f"<b>ОПИСАНИЕ:</b>\n\n{escape_html(preview)}\n\n"
        "━━━━━━━━━━━━━━\n\n"
        "<i>Полное описание сохранено в категории</i>"
    )
    
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("📂 К категории", callback_data=f"open_category_{category_id}")
    )
    
    try:
        bot.edit_message_text(
            text,
            call.message.chat.id,
            call.message.message_id,
            reply_markup=markup,
            parse_mode='HTML'
        )
    except:
        bot.send_message(
            call.message.chat.id,
            text,
            reply_markup=markup,
            parse_mode='HTML'
        )


# ═══════════════════════════════════════════════════════════════
# ГЕНЕРАЦИЯ ИЗОБРАЖЕНИЯ
# ═══════════════════════════════════════════════════════════════

@bot.callback_query_handler(func=lambda call: call.data.startswith("gen_image_"))
def handle_generate_image(call):
    """Генерация изображения с помощью Nano Banana Pro"""
    from ai.image_generator import generate_image
    
    category_id = int(call.data.split("_")[-1])
    user_id = call.from_user.id
    
    # Получаем категорию
    category = db.get_category(category_id)
    if not category:
        safe_answer_callback(bot, call.id, "❌ Категория не найдена")
        return
    
    # Проверяем баланс (30 токенов за изображение)
    user = db.get_user(user_id)
    tokens = user.get('tokens', 0)
    cost = 30
    
    if str(user_id) != str(ADMIN_ID) and tokens < cost:
        bot.answer_callback_query(
            call.id,
            f"❌ Недостаточно токенов. Нужно: {cost} 💎",
            show_alert=True
        )
        return
    
    # Показываем прогресс
    try:
        bot.edit_message_text(
            "🎨 <b>Nano Banana Pro генерирует изображение...</b>\n\n"
            "⏳ Это может занять 10-30 секунд...",
            call.message.chat.id,
            call.message.message_id,
            parse_mode='HTML'
        )
    except:
        pass
    
    safe_answer_callback(bot, call.id, "🚀 Начинаем генерацию...")
    
    # Формируем промпт
    category_name = category['name']
    description = category.get('description', '')
    
    # Создаем промпт на английском
    if description:
        prompt = f"{category_name}, {description[:100]}"
    else:
        prompt = f"{category_name}, professional product photography"
    
    # Генерируем изображение
    result = generate_image(prompt=prompt, aspect_ratio="1:1")
    
    if not result['success']:
        bot.send_message(
            call.message.chat.id,
            f"❌ <b>Ошибка генерации</b>\n\n{result.get('error', 'Неизвестная ошибка')}",
            parse_mode='HTML'
        )
        return
    
    image_bytes = result['image_bytes']
    
    # Списываем токены
    if str(user_id) != str(ADMIN_ID):
        db.update_tokens(user_id, -cost)
        
        # Логируем расход
        db.cursor.execute("""
            INSERT INTO token_expenses (user_id, amount, action, category_id, created_at)
            VALUES (%s, %s, %s, %s, NOW())
        """, (user_id, cost, 'image_generation', category_id))
        db.conn.commit()
    
    # Сохраняем изображение (TODO: в будущем сохранять в БД или на диск)
    
    # Отправляем изображение пользователю
    text = (
        "✅ <b>ИЗОБРАЖЕНИЕ СГЕНЕРИРОВАНО!</b>\n"
        "━━━━━━━━━━━━━━\n\n"
        f"📂 Категория: <b>{escape_html(category_name)}</b>\n"
        f"🎨 Модель: Nano Banana Pro\n"
        f"💎 Списано токенов: <b>{cost}</b>\n\n"
        "Изображение готово к использованию!"
    )
    
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("🔄 Сгенерировать ещё", callback_data=f"gen_image_{category_id}"),
        types.InlineKeyboardButton("📂 К категории", callback_data=f"open_category_{category_id}")
    )
    
    try:
        # Удаляем прогресс
        bot.delete_message(call.message.chat.id, call.message.message_id)
    except:
        pass
    
    # Отправляем изображение
    bot.send_photo(
        call.message.chat.id,
        photo=image_bytes,
        caption=text,
        reply_markup=markup,
        parse_mode='HTML'
    )


print("✅ handlers/category_sections.py загружен")


# ═══════════════════════════════════════════════════════════════
# ПРОСМОТР ГАЛЕРЕИ
# ═══════════════════════════════════════════════════════════════

@bot.callback_query_handler(func=lambda call: call.data.startswith("view_gallery_"))
def handle_view_gallery(call):
    """Просмотр галереи медиа"""
    category_id = int(call.data.split("_")[-1])
    
    category = db.get_category(category_id)
    if not category:
        safe_answer_callback(bot, call.id, "❌ Категория не найдена")
        return
    
    media = category.get('media', [])
    if not media or not isinstance(media, list):
        safe_answer_callback(bot, call.id, "❌ Нет медиа-файлов")
        return
    
    safe_answer_callback(bot, call.id, "📤 Отправляю файлы...")
    
    # Удаляем сообщение с меню
    try:
        bot.delete_message(call.message.chat.id, call.message.message_id)
    except:
        pass
    
    # Отправляем все медиа
    for idx, item in enumerate(media, 1):
        try:
            media_type = item.get('type')
            file_id = item.get('file_id')
            
            if media_type == 'photo':
                bot.send_photo(
                    call.message.chat.id,
                    file_id,
                    caption=f"📸 Изображение {idx}/{len(media)}"
                )
            elif media_type == 'video':
                bot.send_video(
                    call.message.chat.id,
                    file_id,
                    caption=f"🎥 Видео {idx}/{len(media)}"
                )
            elif media_type == 'document':
                file_name = item.get('file_name', 'document')
                bot.send_document(
                    call.message.chat.id,
                    file_id,
                    caption=f"📄 Документ {idx}/{len(media)}: {file_name}"
                )
        except Exception as e:
            print(f"❌ Ошибка отправки медиа: {e}")
    
    # Отправляем меню возврата
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("📂 К медиа", callback_data=f"category_media_{category_id}")
    )
    
    bot.send_message(
        call.message.chat.id,
        f"✅ Показано файлов: {len(media)}",
        reply_markup=markup
    )


@bot.callback_query_handler(func=lambda call: call.data.startswith("clear_gallery_"))
def handle_clear_gallery(call):
    """Очистка галереи медиа"""
    category_id = int(call.data.split("_")[-1])
    
    text = (
        "⚠️ <b>ОЧИСТКА ГАЛЕРЕИ</b>\n"
        "━━━━━━━━━━━━━━\n\n"
        "Вы уверены, что хотите удалить ВСЕ медиа-файлы из этой категории?\n\n"
        "<i>Это действие необратимо!</i>"
    )
    
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        types.InlineKeyboardButton("✅ Да, удалить", callback_data=f"confirm_clear_{category_id}"),
        types.InlineKeyboardButton("❌ Отмена", callback_data=f"category_media_{category_id}")
    )
    
    try:
        bot.edit_message_text(
            text,
            call.message.chat.id,
            call.message.message_id,
            reply_markup=markup,
            parse_mode='HTML'
        )
    except:
        bot.send_message(
            call.message.chat.id,
            text,
            reply_markup=markup,
            parse_mode='HTML'
        )
    
    safe_answer_callback(bot, call.id)


@bot.callback_query_handler(func=lambda call: call.data.startswith("confirm_clear_"))
def handle_confirm_clear_gallery(call):
    """Подтверждение очистки галереи"""
    import json
    
    category_id = int(call.data.split("_")[-1])
    
    # Очищаем медиа
    db.cursor.execute("""
        UPDATE categories 
        SET media = %s::jsonb
        WHERE id = %s
    """, (json.dumps([]), category_id))
    db.conn.commit()
    
    safe_answer_callback(bot, call.id, "✅ Галерея очищена")
    
    # Возвращаем к медиа
    fake_call = type('obj', (object,), {
        'data': f'category_media_{category_id}',
        'from_user': call.from_user,
        'message': call.message,
        'id': call.id
    })()
    
    handle_category_media(fake_call)


# ═══════════════════════════════════════════════════════════════
# ПРАЙС-ЛИСТЫ (EXCEL)
# ═══════════════════════════════════════════════════════════════

@bot.callback_query_handler(func=lambda call: call.data.startswith("download_price_template_"))
def handle_download_price_template(call):
    """Скачать шаблон Excel для прайс-листа"""
    category_id = int(call.data.split("_")[-1])
    
    category = db.get_category(category_id)
    if not category:
        safe_answer_callback(bot, call.id, "❌ Категория не найдена", show_alert=True)
        return
    
    category_name = category.get('name', 'price')
    
    # Создаем Excel шаблон
    import tempfile
    import os
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Создаем рабочую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Прайс-лист"
        
        # Заголовки (универсальные, расширенные)
        headers = ['Название', 'Размеры', 'Вес', 'Цена', 'Единица измерения', 'Описание']
        ws.append(headers)
        
        # Стилизация заголовков
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Устанавливаем ширину столбцов
        ws.column_dimensions['A'].width = 35  # Название
        ws.column_dimensions['B'].width = 20  # Размеры
        ws.column_dimensions['C'].width = 15  # Вес
        ws.column_dimensions['D'].width = 12  # Цена
        ws.column_dimensions['E'].width = 18  # Единица
        ws.column_dimensions['F'].width = 45  # Описание
        
        # Универсальные примеры данных (разные типы товаров/услуг)
        examples = [
            ['Консультация юриста', '-', '-', '3000', 'час', 'Консультация по гражданским делам'],
            ['Смартфон Galaxy S23', '146x71x8 мм', '168 г', '45000', 'шт', 'Смартфон с 6.1" экраном'],
            ['Профиль алюминиевый', '2000x50x30 мм', '0.8 кг/м', '450', 'м.п.', 'Профиль для конструкций'],
            ['Кукла Барби', '30x10x5 см', '200 г', '1200', 'шт', 'Игрушка для детей от 3 лет'],
            ['Доставка по городу', '-', 'до 50 кг', '500', 'заказ', 'Доставка в пределах города']
        ]
        
        for example in examples:
            ws.append(example)
        
        # Добавляем пустые строки для заполнения
        for _ in range(15):
            ws.append(['', '', '', '', '', ''])
        
        # Сохраняем во временный файл
        fd, filepath = tempfile.mkstemp(suffix='.xlsx', prefix='price_template_')
        os.close(fd)
        wb.save(filepath)
        
        # Отправляем файл
        with open(filepath, 'rb') as f:
            filename = f"price_template_{category_name.replace(' ', '_')}.xlsx"
            bot.send_document(
                call.message.chat.id,
                f,
                caption=(
                    f"📊 <b>Универсальный шаблон прайс-листа</b>\n"
                    f"📂 Категория: {escape_html(category_name)}\n\n"
                    "📝 <b>Инструкция:</b>\n"
                    "1. Откройте файл в Excel\n"
                    "2. Заполните свои товары/услуги\n"
                    "3. Удалите примеры данных\n"
                    "4. Можете добавлять свои столбцы\n"
                    "5. Загрузите обратно в бот\n\n"
                    "💡 <b>Совет:</b> Все столбцы и данные будут сохранены"
                ),
                parse_mode='HTML',
                visible_file_name=filename
            )
        
        # Удаляем временный файл
        os.unlink(filepath)
        
        safe_answer_callback(bot, call.id, "✅ Шаблон отправлен")
        
    except ImportError:
        bot.answer_callback_query(
            call.id, 
            "❌ Библиотека openpyxl не установлена. Установите: pip install openpyxl", 
            show_alert=True
        )
    except Exception as e:
        print(f"❌ Ошибка создания шаблона: {e}")
        import traceback
        traceback.print_exc()
        safe_answer_callback(bot, call.id, "❌ Ошибка создания шаблона", show_alert=True)


@bot.callback_query_handler(func=lambda call: call.data.startswith("download_current_price_"))
def handle_download_current_price(call):
    """Скачать текущий прайс-лист"""
    category_id = int(call.data.split("_")[-1])
    
    category = db.get_category(category_id)
    if not category:
        safe_answer_callback(bot, call.id, "❌ Категория не найдена", show_alert=True)
        return
    
    prices_data = category.get('prices', {})
    if not prices_data or not isinstance(prices_data, dict):
        safe_answer_callback(bot, call.id, "❌ Прайс-лист пуст", show_alert=True)
        return
    
    headers = prices_data.get('headers', [])
    rows = prices_data.get('rows', [])
    
    if not headers or not rows:
        safe_answer_callback(bot, call.id, "❌ Нет данных для экспорта", show_alert=True)
        return
    
    category_name = category.get('name', 'price')
    
    # Создаем Excel файл
    import tempfile
    import os
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Прайс-лист"
        
        # Заголовки
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[cell.column_letter].width = 20
        
        # Данные
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Сохраняем
        fd, filepath = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(filepath)
        
        # Отправляем
        with open(filepath, 'rb') as file:
            bot.send_document(
                call.message.chat.id,
                file,
                caption=f"📊 <b>Прайс-лист:</b> {category_name}\n\n"
                        f"📦 Позиций: {len(rows)}\n"
                        f"📋 Колонок: {len(headers)}",
                visible_file_name=f"{category_name}_price.xlsx",
                parse_mode='HTML'
            )
        
        os.unlink(filepath)
        safe_answer_callback(bot, call.id, "✅ Прайс-лист отправлен")
        
    except ImportError:
        bot.answer_callback_query(
            call.id,
            "❌ Библиотека openpyxl не установлена",
            show_alert=True
        )
    except Exception as e:
        print(f"❌ Ошибка экспорта прайса: {e}")
        import traceback
        traceback.print_exc()
        safe_answer_callback(bot, call.id, "❌ Ошибка экспорта", show_alert=True)


@bot.callback_query_handler(func=lambda call: call.data.startswith("delete_price_"))
def handle_delete_price(call):
    """Удаление прайс-листа"""
    category_id = int(call.data.split("_")[-1])
    
    category = db.get_category(category_id)
    if not category:
        safe_answer_callback(bot, call.id, "❌ Категория не найдена", show_alert=True)
        return
    
    # Подтверждение
    text = (
        "⚠️ <b>УДАЛЕНИЕ ПРАЙС-ЛИСТА</b>\n\n"
        "Вы уверены что хотите удалить прайс-лист?\n\n"
        "Это действие нельзя отменить!"
    )
    
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        types.InlineKeyboardButton("✅ Да, удалить", callback_data=f"confirm_delete_price_{category_id}"),
        types.InlineKeyboardButton("❌ Отмена", callback_data=f"category_prices_{category_id}")
    )
    
    try:
        bot.edit_message_text(
            text,
            call.message.chat.id,
            call.message.message_id,
            reply_markup=markup,
            parse_mode='HTML'
        )
    except:
        bot.send_message(call.message.chat.id, text, reply_markup=markup, parse_mode='HTML')
    
    safe_answer_callback(bot, call.id)


@bot.callback_query_handler(func=lambda call: call.data.startswith("confirm_delete_price_"))
def confirm_delete_price(call):
    """Подтверждение удаления прайса"""
    category_id = int(call.data.split("_")[-1])
    
    # Удаляем прайс
    db.cursor.execute("""
        UPDATE categories 
        SET prices = NULL
        WHERE id = %s
    """, (category_id,))
    db.conn.commit()
    
    safe_answer_callback(bot, call.id, "✅ Прайс-лист удален")
    
    # Возвращаемся в меню цен
    call.data = f"category_prices_{category_id}"
    handle_category_prices(call)


# Состояние загрузки прайса
price_upload_state = {}

@bot.callback_query_handler(func=lambda call: call.data.startswith("upload_price_file_"))
def handle_upload_price_file(call):
    """Инструкция по загрузке прайс-листа"""
    category_id = int(call.data.split("_")[-1])
    user_id = call.from_user.id
    
    category = db.get_category(category_id)
    if not category:
        safe_answer_callback(bot, call.id, "❌ Категория не найдена", show_alert=True)
        return
    
    category_name = category.get('name', 'Без названия')
    
    # Устанавливаем состояние ожидания файла
    price_upload_state[user_id] = {
        'category_id': category_id,
        'waiting_file': True
    }
    
    text = (
        f"📤 <b>ЗАГРУЗКА ПРАЙС-ЛИСТА</b>\n"
        f"📂 Категория: {escape_html(category_name)}\n"
        "━━━━━━━━━━━━━━\n\n"
        "📝 <b>Требования к файлу:</b>\n"
        "• Формат: Excel (.xlsx или .xls)\n"
        "• Первая строка - заголовки столбцов\n"
        "• Данные начинаются со 2-й строки\n\n"
        "✅ <b>Все столбцы и строки будут сохранены!</b>\n"
        "Вы можете использовать любую структуру:\n"
        "• Стандартную (из шаблона)\n"
        "• Свою собственную\n"
        "• Добавлять любые столбцы\n\n"
        "💡 <b>Совет:</b> Скачайте шаблон как пример\n\n"
        "👇 Отправьте Excel файл с прайс-листом:"
    )
    
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("❌ Отмена", callback_data=f"category_prices_{category_id}")
    )
    
    try:
        bot.edit_message_text(
            text,
            call.message.chat.id,
            call.message.message_id,
            reply_markup=markup,
            parse_mode='HTML'
        )
    except:
        bot.send_message(call.message.chat.id, text, reply_markup=markup, parse_mode='HTML')
    
    safe_answer_callback(bot, call.id)


@bot.message_handler(content_types=['document'], func=lambda message: message.from_user.id in price_upload_state 
                     and price_upload_state[message.from_user.id].get('waiting_file'))
def handle_price_file_upload(message):
    """Обработка загруженного Excel файла с прайс-листом"""
    user_id = message.from_user.id
    
    state = price_upload_state.get(user_id)
    if not state:
        return
    
    category_id = state['category_id']
    
    # Проверяем что это Excel файл
    if not (message.document.file_name.endswith('.xlsx') or message.document.file_name.endswith('.xls')):
        bot.send_message(
            message.chat.id,
            "❌ Пожалуйста, отправьте Excel файл (.xlsx или .xls)"
        )
        return
    
    try:
        from openpyxl import load_workbook
        import tempfile
        import os
        
        # Скачиваем файл
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        
        # Сохраняем во временный файл
        fd, filepath = tempfile.mkstemp(suffix='.xlsx')
        with os.fdopen(fd, 'wb') as f:
            f.write(downloaded_file)
        
        # Читаем Excel
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # Читаем заголовки (первая строка)
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        if not headers:
            bot.send_message(
                message.chat.id,
                "❌ Не найдены заголовки таблицы (первая строка должна содержать названия столбцов)"
            )
            os.unlink(filepath)
            return
        
        # Парсим ВСЕ данные (все столбцы, все строки) с сохранением форматирования
        prices_data = []
        for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
            # Пропускаем полностью пустые строки
            if not any(cell.value for cell in row_cells):
                continue
            
            # Создаем словарь со ВСЕМИ столбцами
            row_data = {}
            for col_idx, cell in enumerate(row_cells):
                if col_idx < len(headers):
                    header_name = headers[col_idx]
                    
                    # Получаем значение ячейки
                    cell_value = cell.value
                    
                    # Если ячейка пустая - пустая строка
                    if cell_value is None:
                        row_data[header_name] = ''
                        continue
                    
                    # Если число - проверяем формат
                    if isinstance(cell_value, (int, float)):
                        # Проверяем есть ли формат валюты
                        if cell.number_format and ('₽' in cell.number_format or '₽' in str(cell.number_format)):
                            row_data[header_name] = f"{cell_value} ₽"
                        elif cell.number_format and ('руб' in cell.number_format.lower()):
                            row_data[header_name] = f"{cell_value} руб."
                        else:
                            row_data[header_name] = str(cell_value).strip()
                    else:
                        # Строковое значение
                        row_data[header_name] = str(cell_value).strip()
            
            # Добавляем строку если есть хотя бы одно непустое значение
            if any(row_data.values()):
                prices_data.append(row_data)
        
        # Удаляем временный файл
        os.unlink(filepath)
        
        if not prices_data:
            bot.send_message(
                message.chat.id,
                "❌ В файле не найдено данных"
            )
            return
        
        print(f"📊 Считано {len(prices_data)} строк из Excel:")
        for i, row in enumerate(prices_data[:3], 1):  # Первые 3 строки для примера
            print(f"   Строка {i}: {row}")
        
        # Сохраняем в БД (со всеми столбцами и данными)
        import json
        save_data = {
            'headers': headers,
            'rows': prices_data
        }
        
        db.cursor.execute("""
            UPDATE categories 
            SET prices = %s::jsonb
            WHERE id = %s
        """, (json.dumps(save_data, ensure_ascii=False), category_id))
        db.conn.commit()
        
        # Очищаем состояние
        del price_upload_state[user_id]
        
        text = (
            "✅ <b>ПРАЙС-ЛИСТ ЗАГРУЖЕН</b>\n"
            "━━━━━━━━━━━━━━\n\n"
            f"📊 Столбцов: <b>{len(headers)}</b>\n"
            f"📊 Строк: <b>{len(prices_data)}</b>\n\n"
            f"📋 Столбцы: {', '.join(headers[:3])}"
        )
        
        if len(headers) > 3:
            text += f", <i>+{len(headers) - 3} ещё</i>"
        
        text += "\n\n<b>Первые 3 строки:</b>\n"
        
        for i, row_data in enumerate(prices_data[:3], 1):
            # Показываем первые 3 столбца каждой строки
            row_values = []
            for header in headers[:3]:
                value = row_data.get(header, '')
                if value:
                    row_values.append(escape_html(value))
            
            if row_values:
                text += f"{i}. {' | '.join(row_values)}\n"
        
        if len(prices_data) > 3:
            text += f"\n<i>... и ещё {len(prices_data) - 3} строк</i>"
        
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("🔙 К ценам", callback_data=f"category_prices_{category_id}")
        )
        
        bot.send_message(message.chat.id, text, reply_markup=markup, parse_mode='HTML')
        
    except ImportError:
        bot.send_message(
            message.chat.id,
            "❌ Библиотека openpyxl не установлена. Установите: pip install openpyxl --break-system-packages"
        )
    except Exception as e:
        print(f"❌ Ошибка обработки файла: {e}")
        import traceback
        traceback.print_exc()
        
        bot.send_message(
            message.chat.id,
            f"❌ Ошибка обработки файла: {e}"
        )

